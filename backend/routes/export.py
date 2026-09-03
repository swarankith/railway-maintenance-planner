"""
Export API Endpoints:
- GET /api/v1/export/requests (Excel & PDF)
- GET /api/v1/export/approvals (Excel & PDF)
Generates government-grade formatted Excel and PDF reports.
"""
import io
from datetime import datetime
from typing import Optional, List
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from backend.config import APP_TIMEZONE
from backend.database import get_db
from backend.models import DBMaintenanceRequest, DBApprovalAudit, DBUser
from backend.auth import get_current_user

router = APIRouter(prefix="/api/v1/export", tags=["Exports"])


def get_timestamp_filename(prefix: str, extension: str) -> str:
    now_str = datetime.now(APP_TIMEZONE).strftime("%Y-%m-%d_%H-%M-%S")
    return f"{prefix}_{now_str}.{extension}"


@router.get("/requests")
def export_requests(
    format: str = Query("excel", pattern="^(excel|xlsx|pdf)$"),
    corridor: Optional[str] = None,
    department: Optional[str] = None,
    status: Optional[str] = None,
    current_user: DBUser = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Exports all maintenance requests with application IDs, status, priorities, and reasons to Excel or PDF.
    """
    query = db.query(DBMaintenanceRequest)
    if corridor:
        query = query.filter(DBMaintenanceRequest.corridor.ilike(f"%{corridor}%"))
    if department:
        query = query.filter(DBMaintenanceRequest.department == department)
    if status:
        query = query.filter(DBMaintenanceRequest.status == status)

    records = query.order_by(DBMaintenanceRequest.priority.asc(), DBMaintenanceRequest.earliest_start.asc()).all()

    filename = get_timestamp_filename("requests_status", "xlsx" if format in ("excel", "xlsx") else "pdf")

    if format in ("excel", "xlsx"):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Maintenance Requests"

        # Headers
        headers = [
            "Application ID", "Request ID", "Department", "Corridor", "Work Type",
            "Priority", "Block Type", "Start (IST)", "End (IST)", "KM Range",
            "Status", "Resources", "Validation Notes"
        ]
        ws.append(headers)

        # Style header row (Navy Blue with White text)
        header_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Data rows
        for row_idx, r in enumerate(records, start=2):
            start_str = r.earliest_start.strftime("%Y-%m-%d %H:%M") if r.earliest_start else ""
            end_str = r.latest_end.strftime("%Y-%m-%d %H:%M") if r.latest_end else ""
            km_str = f"KM {r.km_start:.1f}-{r.km_end:.1f}"
            res_str = ", ".join(r.required_resources or [])

            prio_label = f"P{r.priority}"
            if r.priority == 1:
                prio_label = "P1 - Emergency"
            elif r.priority == 2:
                prio_label = "P2 - High Urgent"
            elif r.priority == 3:
                prio_label = "P3 - Normal"

            row_data = [
                r.application_id or "APP-LEGACY",
                r.request_id,
                r.department,
                r.corridor,
                r.work_type,
                prio_label,
                r.block_type,
                start_str,
                end_str,
                km_str,
                r.status,
                res_str,
                r.validation_notes or ""
            ]
            ws.append(row_data)

            # Alternate row coloring
            if row_idx % 2 == 0:
                row_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = row_fill

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    else:
        # PDF Generation via ReportLab
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(letter), leftMargin=20, rightMargin=20, topMargin=20, bottomMargin=20)
        elements = []

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            name="TitleStyle",
            parent=styles["Heading1"],
            fontSize=16,
            textColor=colors.HexColor("#000080"),
            spaceAfter=10
        )
        elements.append(Paragraph("Indian Railways — Maintenance Requests Status Report", title_style))
        elements.append(Paragraph(f"Generated: {datetime.now(APP_TIMEZONE).strftime('%d-%b-%Y %H:%M:%S IST')} | User: {current_user.username} ({current_user.role})", styles["Normal"]))
        elements.append(Spacer(1, 12))

        table_data = [[
            "App ID", "Req ID", "Dept", "Corridor", "Work Type", "Prio", "Span", "Status", "Validation"
        ]]

        for r in records:
            km_str = f"{r.km_start:.1f}-{r.km_end:.1f}"
            table_data.append([
                Paragraph((r.application_id or "")[:15], styles["Normal"]),
                Paragraph(r.request_id, styles["Normal"]),
                Paragraph(r.department, styles["Normal"]),
                Paragraph(r.corridor, styles["Normal"]),
                Paragraph(r.work_type[:24], styles["Normal"]),
                f"P{r.priority}",
                km_str,
                r.status,
                Paragraph((r.validation_notes or "")[:35], styles["Normal"])
            ])

        pdf_table = Table(table_data, colWidths=[80, 60, 65, 80, 130, 40, 55, 65, 180])
        pdf_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#000080')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F8FAFC')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(pdf_table)
        doc.build(elements)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )


@router.get("/approvals")
def export_approvals(
    format: str = Query("excel", pattern="^(excel|xlsx|pdf)$"),
    schedule_id: Optional[str] = None,
    current_user: DBUser = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Exports full approval audit history to Excel or PDF.
    """
    query = db.query(DBApprovalAudit)
    if schedule_id:
        query = query.filter(DBApprovalAudit.schedule_id == schedule_id)

    audits = query.order_by(DBApprovalAudit.timestamp.desc()).all()
    filename = get_timestamp_filename("approval_history", "xlsx" if format in ("excel", "xlsx") else "pdf")

    if format in ("excel", "xlsx"):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Approval History"

        headers = ["Audit ID", "Schedule ID", "Application ID", "Action", "Role", "User Name", "Timestamp (IST)", "Reason / Notes"]
        ws.append(headers)

        header_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_idx, a in enumerate(audits, start=2):
            ts_str = a.timestamp.strftime("%Y-%m-%d %H:%M:%S IST") if a.timestamp else ""
            row_data = [
                a.id,
                a.schedule_id,
                a.application_id or "APP-GLOBAL",
                a.action,
                a.role,
                a.user_name,
                ts_str,
                a.notes or ""
            ]
            ws.append(row_data)

            if row_idx % 2 == 0:
                row_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = row_fill

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 14)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    else:
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(letter), leftMargin=20, rightMargin=20, topMargin=20, bottomMargin=20)
        elements = []

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            name="TitleStyle",
            parent=styles["Heading1"],
            fontSize=16,
            textColor=colors.HexColor("#000080"),
            spaceAfter=10
        )
        elements.append(Paragraph("Indian Railways — Approval & Decision Audit History", title_style))
        elements.append(Paragraph(f"Exported: {datetime.now(APP_TIMEZONE).strftime('%d-%b-%Y %H:%M:%S IST')} | User: {current_user.username}", styles["Normal"]))
        elements.append(Spacer(1, 12))

        table_data = [[
            "Audit ID", "Schedule ID", "Action", "Role", "User Name", "Timestamp (IST)", "Notes / Reason"
        ]]

        for a in audits:
            ts_str = a.timestamp.strftime("%d-%b %H:%M") if a.timestamp else ""
            table_data.append([
                str(a.id),
                a.schedule_id,
                a.action,
                a.role,
                a.user_name,
                ts_str,
                Paragraph(a.notes or "", styles["Normal"])
            ])

        pdf_table = Table(table_data, colWidths=[50, 100, 70, 90, 90, 80, 270])
        pdf_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#000080')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F8FAFC')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(pdf_table)
        doc.build(elements)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
